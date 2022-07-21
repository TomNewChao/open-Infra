# -*- coding: utf-8 -*-
# @Time    : 2022/7/7 10:30
# @Author  : Tom_zc
# @FileName: scan_port.py
# @Software: PyCharm
import os
import re

import requests
import argparse
import yaml
import subprocess
import openpyxl
from abc import abstractmethod
from .common import func_retry
from collections import defaultdict

from requests.packages.urllib3.exceptions import InsecureRequestWarning
from huaweicloudsdkcore.auth.credentials import BasicCredentials
from huaweicloudsdkcore.http.http_config import HttpConfig
from huaweicloudsdkcore.client import Client
from huaweicloudsdkeip.v2 import EipClient as EipClientV2
from huaweicloudsdkeip.v2 import ListPublicipsRequest as ListPublicipsRequestV2
from huaweicloudsdkeip.v3 import EipClient as EipClientV3
from huaweicloudsdkeip.v3 import ListPublicipsRequest as ListPublicipsRequestV3

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


# noinspection DuplicatedCode
class GlobalConfig(object):
    base_path = os.path.dirname(__file__)
    txt_path = os.path.join(base_path, "ip.txt")

    ip_result_path = os.path.join(base_path, "ip_result.txt")
    config_path = os.path.join(base_path, "scan_port.yaml")
    zone_alias_dict = {
        "cn-north-1": "华北-北京一",
        "cn-north-4": "华北-北京四",
        "cn-north-5": "华北-乌兰察布二零一",
        "cn-north-6": "华北-乌兰察布二零二",
        "cn-north-9": "华北-乌兰察布一",
        "cn-east-3": "华东-上海一",
        "cn-east-2": "华东-上海二",
        "cn-south-1": "华南-广州",
        "cn-south-4": "华南-广州-友好用户环境",
        "cn-southwest-2": "西南-贵阳一",
        "ap-southeast-1": "中国-香港",
        "ap-southeast-2": "亚太-曼谷",
        "ap-southeast-3": "亚太-新加坡",
        "af-south-1": "非洲-约翰内斯堡",
        "na-mexico-1": "拉美-墨西哥城一",
        "la-north-2": "拉美-墨西哥城二",
        "sa-brazil-1": "拉美-圣保罗一",
        "la-south-2": "拉美-圣地亚哥",
        "ru-northwest-2": "俄罗斯-莫斯科二",
    }
    eip_v2_zone = ["cn-south-4", ]

    excel_path = os.path.join(base_path, "公网IP端口扫描统计表.xlsx")
    excel_title = ["弹性公网IP", "端口", "状态", "链接协议", "传输协议"]
    excel_server_info_title = ["弹性公网IP", "端口", "服务器版本信息"]
    need_delete_sheet_name = "Sheet"

    tcp_search_cmd = "nmap -sS -Pn -n --open --min-hostgroup 4 --min-parallelism 1024 --host-timeout 180 -T4 -v -oG ip_result.txt {}"
    udp_search_cmd = "nmap -sU --min-hostgroup 4 --min-parallelism 1024 --host-timeout 180 -v -oG ip_result.txt {}"


# noinspection DuplicatedCode
class EndPoint(object):
    vpc_endpoint = "https://vpc.{}.myhuaweicloud.com"
    nat_endpoint = "https://nat.{}.myhuaweicloud.com"
    elb_endpoint = "https://elb.{}.myhuaweicloud.com"
    bms_endpoint = "https://bms.{}.myhuaweicloud.com"
    ecs_endpoint = "https://ecs.{}.myhuaweicloud.com"
    rds_endpoint = "https://rds.{}.myhuaweicloud.com"


# noinspection PyUnresolvedReferences
class BaseInstance(object):
    def __init__(self, base_client, config, credentials, endpoint):
        if not issubclass(base_client, Client):
            raise Exception("base client must be client")
        self.base_client = base_client.new_builder() \
            .with_http_config(config) \
            .with_credentials(credentials) \
            .with_endpoint(endpoint) \
            .build()

    @abstractmethod
    def set_req_method(self):
        pass

    @abstractmethod
    def parse_response_data(self, response_dict):
        pass

    def show_infos(self, *args, **kwargs):
        info_request, method = self.set_req_method()
        show_infos_req = info_request(*args, **kwargs)
        show_infos_method = getattr(self.base_client, method)
        ret = show_infos_method(show_infos_req)
        return ret.to_dict()


class EipInstanceV2(BaseInstance):
    def __init__(self, *args, **kwargs):
        super(EipInstanceV2, self).__init__(*args, **kwargs)

    def set_req_method(self):
        return ListPublicipsRequestV2, "list_publicips"

    def parse_response_data(self, response_dict):
        return response_dict['publicips']


class EipInstanceV3(BaseInstance):
    def __init__(self, *args, **kwargs):
        super(EipInstanceV3, self).__init__(*args, **kwargs)

    def set_req_method(self):
        return ListPublicipsRequestV3, "list_publicips"

    def parse_response_data(self, response_dict):
        return response_dict['publicips']


# noinspection DuplicatedCode
class EipTools(object):
    def __init__(self, *args, **kwargs):
        super(EipTools, self).__init__(*args, **kwargs)

    @classmethod
    def get_eip_config(cls):
        config = HttpConfig.get_default_config()
        config.ignore_ssl_verification = True
        return config

    @classmethod
    def output_txt(cls, eip_info_list):
        with open(GlobalConfig.txt_path, "w") as f:
            for content in eip_info_list:
                f.write(content)
                f.write("\n")

    @classmethod
    def read_ip_txt(cls):
        with open(GlobalConfig.txt_path, "r") as f:
            return f.readlines()

    @classmethod
    def read_ip_result_txt(cls):
        with open(GlobalConfig.ip_result_path, "r") as f:
            content = f.readlines()
        return content

    # noinspection PyUnresolvedReferences
    @classmethod
    def parse_tcp_result_txt_all(cls, tcp_content_list):
        host_data = defaultdict(list)
        for info in tcp_content_list:
            if "Host:" in info and "Ports:":
                ret_list = list()
                info_list = info.split("Ports:")
                ip = re.match(r"Host: (.*?)\(\)", info_list[0])
                if not ip:
                    continue
                ip = ip.groups()[0].strip()
                infor_str_list = info_list[1].split()
                for infor_str_temp in infor_str_list:
                    if "/" in infor_str_temp:
                        for infor_str_temp_data in infor_str_temp.split(","):
                            port = re.match(r"(.*?)///", infor_str_temp_data.strip())
                            if not port:
                                continue
                            ret_list.append(port.groups()[0].strip())
                host_data[ip] = ret_list
        return host_data

    @classmethod
    def parse_tcp_result_txt(cls, tcp_content_list):
        ret_list = list()
        for info in tcp_content_list:
            if "Host:" in info and "Ports:" in info:
                info_list = info.split("Ports:")
                print(info_list)
                infor_str_list = info_list[1].split()
                for infor_str_temp in infor_str_list:
                    if "/" in infor_str_temp:
                        for infor_str_temp_data in infor_str_temp.split(","):
                            port = re.match(r"(.*?)///", infor_str_temp_data.strip())
                            if not port:
                                continue
                            port_str = port.groups()[0].strip()
                            port_content = list(filter(lambda x: x != "", port_str.split('/')))
                            ret_list.append(port_content)
        return ret_list

    @classmethod
    def get_device_info(cls, instance_list):
        ret_dict = dict()
        for instance_temp in instance_list:
            instance_info = instance_temp.show_infos()
            device_info = instance_temp.parse_response_data(instance_info)
            for key, value in device_info.items():
                if key not in ret_dict.keys():
                    ret_dict[key] = value
        return ret_dict

    @classmethod
    def parse_input_args(cls):
        par = argparse.ArgumentParser()
        par.add_argument("-config_path", "--config_path", help="The config path of object", required=False)
        args = par.parse_args()
        return args

    @staticmethod
    def load_yaml(file_path, method="load"):
        """
        method: load_all/load
        """
        yaml_load_method = getattr(yaml, method)
        with open(file_path, "r", encoding="utf-8") as file:
            content = yaml_load_method(file, Loader=yaml.FullLoader)
        return content

    @classmethod
    def check_config_data(cls, config_list):
        for config_temp in config_list:
            if config_temp.get("account") is None:
                raise Exception("Account is invalid")
            if config_temp.get("ak") is None:
                raise Exception("Ak is invalid")
            if config_temp.get("sk") is None:
                raise Exception("Sk is invalid")
            if not isinstance(config_temp.get("project_info"), list):
                raise Exception("project info must be dict")
            for project_temp in config_temp["project_info"]:
                if project_temp.get("project_id") is None:
                    raise Exception("project_id is invalid")
                if project_temp.get("zone") is None:
                    raise Exception("project_id is invalid")

    @func_retry(tries=1)
    def get_data_list(self, eip_tools, project_temp, ak, sk):
        project_id = project_temp["project_id"]
        zone = project_temp["zone"]
        config = eip_tools.get_eip_config()
        credentials = BasicCredentials(ak, sk, project_id)
        if zone in GlobalConfig.eip_v2_zone:
            eip_instance = EipInstanceV2(EipClientV2, config, credentials, EndPoint.vpc_endpoint.format(zone))
        else:
            eip_instance = EipInstanceV3(EipClientV3, config, credentials, EndPoint.vpc_endpoint.format(zone))
        eip_dict = eip_instance.show_infos()
        eip_list = eip_instance.parse_response_data(eip_dict)
        eip_ip_list = list()
        for eip_info in eip_list:
            eip_ip_list.append(eip_info['public_ip_address'])
        return eip_ip_list

    @classmethod
    def execute_cmd(cls, cmd):
        """
        Execute commands through subprocess
        :param cmd: string, the cmd
        :return: string, the result of execute cmd
        """
        return subprocess.getoutput(cmd)

    @classmethod
    def output_excel(cls, tcp_dict, username, is_server_info=None):
        if os.path.exists(GlobalConfig.excel_path):
            work_book = openpyxl.load_workbook(GlobalConfig.excel_path)
        else:
            work_book = openpyxl.Workbook()
        if username not in work_book.get_sheet_names():
            work_book.create_sheet(username)
        if GlobalConfig.need_delete_sheet_name in work_book.get_sheet_names():
            need_remove_sheet = work_book.get_sheet_by_name(GlobalConfig.need_delete_sheet_name)
            work_book.remove_sheet(need_remove_sheet)
        table = work_book.get_sheet_by_name(username)
        table.delete_rows(1, 65536)
        if not is_server_info:
            table.append(GlobalConfig.excel_title)
        else:
            table.append(GlobalConfig.excel_server_info_title)
        for ip, eip_info_list in tcp_dict.items():
            for eip_list in eip_info_list:
                if eip_list:
                    temp_info = [ip]
                    temp_info.extend(eip_list)
                    table.append(temp_info)
        work_book.save(GlobalConfig.excel_path)

    @classmethod
    def request_server(cls, ip, ports):
        url = r"http://{}:{}/".format(ip, ports)
        try:
            ret = requests.get(url, timeout=(180, 180))
            server_info = ret.headers.get("Server", "Unknown")
        except Exception as e:
            print("collect url:{}, err:{}".format(url, e))
            server_info = str(e)
        return server_info

    @classmethod
    def collect_tcp_server_info(cls, tcp_ret_dict):
        server_dict = defaultdict(list)
        for ip, ip_info in tcp_ret_dict.items():
            for ip_temp in ip_info:
                server_info = cls.request_server(ip, ip_temp[0])
                server_dict[ip].append([ip_temp[0], server_info])
        return server_dict


# noinspection DuplicatedCode
def main():
    eip_tools = EipTools()
    input_args = eip_tools.parse_input_args()
    print("##################1.start to parse params #############")
    if not input_args.config_path:
        config_path = GlobalConfig.config_path
    else:
        config_path = input_args.config_path
    config_list = eip_tools.load_yaml(config_path)
    eip_tools.check_config_data(config_list)
    print("############2.start to collect and output to excel######")
    result_list = list()
    for config_item in config_list:
        ak = config_item["ak"]
        sk = config_item["sk"]
        project_info = config_item["project_info"]
        for project_temp in project_info:
            print("Collect the zone of info:{}".format(project_temp["zone"]))
            ret_temp = eip_tools.get_data_list(eip_tools, project_temp, ak, sk)
            result_list.extend(ret_temp or [])
    result_list = list(set(result_list))
    print("Write the data to txt, the count of ip:{}...".format(len(result_list)))
    if not result_list:
        return
    # eip_tools.output_txt(result_list)
    # result_list = eip_tools.read_ip_txt()
    print("###########3.lookup port###################")
    tcp_ret_dict, udp_ret_dict = dict(), dict()
    for ip in result_list:
        print("************************")
        # ip = ip_infor.split("\n")[0]
        print("parse ip:{}".format(ip))
        print("1.start to collect tcp info")
        eip_tools.execute_cmd(GlobalConfig.tcp_search_cmd.format(ip))
        tcp_content_list = eip_tools.read_ip_result_txt()
        print("parse tcp content:{}".format(tcp_content_list))
        tcp_temp_list = eip_tools.parse_tcp_result_txt(tcp_content_list)
        print("parse tcp port:{}".format(tcp_temp_list))
        tcp_ret_dict[ip] = tcp_temp_list

        print("2.start to collect udp info")
        eip_tools.execute_cmd(GlobalConfig.udp_search_cmd.format(ip))
        udp_content_list = eip_tools.read_ip_result_txt()
        print("parse udp content:{}".format(udp_content_list))
        udp_temp_list = eip_tools.parse_tcp_result_txt(udp_content_list)
        print("parse udp port:{}".format(udp_temp_list))
        udp_ret_dict[ip] = udp_temp_list
    print("Write the data to excel, the count of tcp ip:{}...".format(len(tcp_ret_dict.keys())))
    print("Write the data to excel, the count of udp ip:{}...".format(len(udp_ret_dict.keys())))
    eip_tools.output_excel(tcp_ret_dict, "tcp")
    eip_tools.output_excel(udp_ret_dict, "udp")
    print("###########4.query nginx server###################")
    tcp_server_info = EipTools.collect_tcp_server_info(tcp_ret_dict)
    eip_tools.output_excel(tcp_server_info, "tcp_server_info", is_server_info=True)
    print("##################5.finish################")


if __name__ == "__main__":
    main()
