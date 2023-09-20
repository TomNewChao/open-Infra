# -*- coding: utf-8 -*-
# @Time    : 2022/7/7 10:30
# @Author  : Tom_zc
# @FileName: common.py
# @Software: PyCharm
import datetime
import time
import base64
import string
import random
import pickle
import subprocess
import tempfile
import traceback
import yaml
import openpyxl

from functools import wraps
from collections import Mapping
from collections import Iterable
from django.conf import settings
from logging import getLogger
from django.http import JsonResponse, HttpResponse
from open_infra.utils.api_error_code import ErrCode

try:
    from StringIO import StringIO
    from urllib import urlencode
except ImportError:
    # Python3
    from io import BytesIO as StringIO
    from urllib.parse import urlencode

logger = getLogger("django")


def func_retry(tries=3, delay=1):
    """the func retry decorator"""
    def deco_retry(fn):
        @wraps(fn)
        def inner(*args, **kwargs):
            for i in range(tries):
                try:
                    return fn(*args, **kwargs)
                except Exception as e:
                    logger.error("func_retry:{} e:{} traceback: {}".format(fn.__name__, e, traceback.format_exc()))
                    time.sleep(delay)
            else:
                logger.info("func_retry:{} over tries, failed".format(fn.__name__))

        return inner

    return deco_retry


def func_catch_exception(fn):
    """the func catch exception fun"""
    @wraps(fn)
    def inner(*args, **kwargs):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            logger.error("func_retry:{} e:{} traceback: {}".format(fn.__name__, e, traceback.format_exc()))
    return inner


def pick_dumps(json_dict):
    """pick dumps"""
    json_bytes = pickle.dumps(json_dict)
    json_secret = base64.b64encode(json_bytes)
    json_str = json_secret.decode()
    return json_str


def pick_loads(json_str):
    """pick loads"""
    json_secret = json_str.encode()
    json_bytes = base64.b64decode(json_secret)
    json_dict = pickle.loads(json_bytes)
    return json_dict


# noinspection PyShadowingBuiltins
def bytes_convert_str(input):
    """convert bytes to str, Can be processed recursively"""
    if isinstance(input, dict):
        return {bytes_convert_str(key): bytes_convert_str(value) for key, value in input.items()}
    elif isinstance(input, (list, tuple)):
        return [bytes_convert_str(element) for element in input]
    elif isinstance(input, bytes):
        return input.decode('utf-8')
    elif isinstance(input, str):
        return input
    elif isinstance(input, Iterable):
        return [bytes_convert_str(element) for element in input]
    else:
        return input


# noinspection PyShadowingBuiltins
def unicode_convert(input):
    """convert bytes to str, Can be processed recursively"""
    if isinstance(input, Mapping):
        return {unicode_convert(key): unicode_convert(value) for key, value in input.items()}
    elif isinstance(input, (tuple, list)):
        return [unicode_convert(element) for element in input]
    elif isinstance(input, bytes):
        return input.decode('utf-8')
    else:
        return input


def execute_cmd3(cmd, timeout=30, err_log=True):
    """execute cmd"""
    try:
        logger.debug("execute_cmd3 call cmd: %s" % cmd)
        p = subprocess.Popen(cmd, stderr=subprocess.PIPE, stdout=subprocess.PIPE, shell=True, close_fds=True)
        t_wait_seconds = 0
        while True:
            if p.poll() is not None:
                break
            if timeout >= 0 and t_wait_seconds >= (timeout * 100):
                p.terminate()
                return -1, "", "execute_cmd3 exceeded time {0} seconds in executing: {1}".format(timeout, cmd)
            time.sleep(0.01)
            t_wait_seconds += 1
        out, err = p.communicate()
        out, err = bytes_convert_str(out), bytes_convert_str(err)
        ret = p.returncode
        if ret != 0 and err_log:
            logger.error("execute_cmd3 cmd %s return %s, std output: %s, err output: %s.", cmd, ret, out, err)

        return ret, out, err
    except Exception as e:
        return -1, "", "execute_cmd3 exceeded raise, e={0}, trace={1}".format(e.args[0], traceback.format_exc())


def execute_cmd3_with_tmp(cmd_str, timeout=30):
    """Execution of commands with excessive output results
    :param cmd_str
    :param timeout
    :return:
    """
    try:
        with tempfile.NamedTemporaryFile() as out_tmp_file:
            cmd = cmd_str + ' > {}'.format(out_tmp_file.name)
            ret, out, err = execute_cmd3(cmd, timeout=timeout)
            # logger.error("--------------------------execute_cmd3_with_tmp0:{}".format(cmd))
            # logger.error("--------------------------execute_cmd3_with_tmp1:{}".format(out))
            # logger.error("--------------------------execute_cmd3_with_tmp2:{}".format(err))
            if ret != 0:
                return ret, out, err
            out_tmp_file.seek(0)
            with open(out_tmp_file.name, 'r') as f:
                out_data = f.read()
            # logger.error("--------------------------execute_cmd3_with_tmp3:{}".format(ret))
            # logger.error("--------------------------execute_cmd3_with_tmp4:{}".format(out_data))
            # logger.error("--------------------------execute_cmd3_with_tmp5:{}".format(err))
            return ret, out_data, err
    except Exception as e:
        return -1, "", "execute_cmd3_with_tmp exceeded raise, e={0}, trace={1}".format(e.args[0],
                                                                                       traceback.format_exc())


def auto_response():
    """
    Decorator that reports the execution time.
    """

    def decorator(func):

        @wraps(func)
        def wrapper(*args, **kw):
            # noinspection PyBroadException
            try:
                data = func(*args, **kw)
            except MgrException as e:
                logger.error(traceback.format_exc())
                return assemble_api_result(e.code, e.trans_code, e.trans_para)
            except Exception:
                logger.error(traceback.format_exc())
                return assemble_api_result(ErrCode.INTERNAL_ERROR)
            if isinstance(data, HttpResponse):
                return data
            else:
                return assemble_api_result(ErrCode.STATUS_SUCCESS, data=data)

        return wrapper

    return decorator


def translate_error_desc(trans_code, trans_para=None):
    """translate return error desc"""
    if trans_para is None:
        trans_para = list()
    tmp_desc = ErrCode.get_err_desc(trans_code)
    if tmp_desc is not None:
        try:
            tmp_desc = unicode_convert(tmp_desc)
            trans_desc = tmp_desc
            if len(trans_para) != 0:
                trans_para = unicode_convert(trans_para)
                trans_desc = tmp_desc % (tuple(trans_para))
        except Exception as e:
            logger.warning("Format error desc for %s failed: %s.", trans_code, str(e))
            trans_desc = tmp_desc
    else:
        if trans_code == 0:
            trans_desc = ErrCode.get_err_desc(trans_code)
        else:
            trans_desc = ErrCode.get_err_desc(ErrCode.STATUS_FAILED)

    return trans_desc


def assemble_api_result(err_code, trans_code=None, trans_para=None, data=None, lang_flag=None, replace_none=True):
    """return api result"""
    if trans_para is None:
        trans_para = []
    else:
        if isinstance(trans_para, str):
            trans_para = [trans_para]
    if replace_none and data is None:
        data = {}
    api_ret = {'code': err_code}
    if not trans_code:
        trans_code = err_code
    tmp_desc = ErrCode.get_err_desc(trans_code, lang_flag=lang_flag)
    if tmp_desc is not None:
        try:
            tmp_desc = unicode_convert(tmp_desc)
            trans_desc = tmp_desc
            if len(trans_para) != 0:
                trans_para = unicode_convert(trans_para)
                trans_desc = tmp_desc % (tuple(trans_para))
        except Exception as e:
            logger.warning("Format error desc for %s failed: %s.", trans_code, str(e))
            trans_desc = tmp_desc
        api_ret['description'] = trans_desc
    else:
        if err_code == 0:
            trans_desc = ErrCode.get_err_desc(err_code, lang_flag=lang_flag)
        else:
            trans_desc = ErrCode.get_err_desc(ErrCode.STATUS_FAILED, lang_flag=lang_flag)

        api_ret['description'] = trans_desc
    api_ret['data'] = data
    return JsonResponse(data=api_ret)


def load_yaml(file_path, method="load"):
    """read yaml to yaml obj"""
    yaml_load_method = getattr(yaml, method)
    with open(file_path, "r", encoding="utf-8") as file:
        return yaml_load_method(file, Loader=yaml.FullLoader)


def convert_yaml(content, method="load"):
    """str to  yaml obj"""
    yaml_load_method = getattr(yaml, method)
    return yaml_load_method(content, Loader=yaml.FullLoader)


def output_scan_port_excel(tcp_info, udp_info):
    """output scan port list data to excel"""
    work_book = openpyxl.Workbook()
    if settings.EXCEL_TCP_PAGE_NAME not in work_book.get_sheet_names():
        work_book.create_sheet(settings.EXCEL_TCP_PAGE_NAME)
    if settings.EXCEL_UDP_PAGE_NAME not in work_book.get_sheet_names():
        work_book.create_sheet(settings.EXCEL_UDP_PAGE_NAME)
    if settings.DEFAULT_SHEET_NAME in work_book.get_sheet_names():
        need_remove_sheet = work_book.get_sheet_by_name(settings.DEFAULT_SHEET_NAME)
        work_book.remove_sheet(need_remove_sheet)
    table = work_book.get_sheet_by_name(settings.EXCEL_TCP_PAGE_NAME)
    table.delete_rows(1, 65536)
    table.append(settings.EXCEL_TITLE)
    for eip_info_list in tcp_info:
        table.append(eip_info_list)

    table = work_book.get_sheet_by_name(settings.EXCEL_UDP_PAGE_NAME)
    table.delete_rows(1, 65536)
    table.append(settings.EXCEL_TITLE)
    for eip_info_list in udp_info:
        table.append(eip_info_list)

    buf = StringIO()
    work_book.save(buf)
    buf.seek(0)
    return buf.read()


def output_scan_obs_excel(anonymous_file_list, anonymous_bucket_list):
    """output scan obs list data to excel"""
    work_book = openpyxl.Workbook()
    if settings.OBS_ANONYMOUS_BUCKET_PAGE_NAME not in work_book.get_sheet_names():
        work_book.create_sheet(settings.OBS_ANONYMOUS_BUCKET_PAGE_NAME)
    if settings.OBS_SENSITIVE_FILE_PAGE_NAME not in work_book.get_sheet_names():
        work_book.create_sheet(settings.OBS_SENSITIVE_FILE_PAGE_NAME)
    if settings.DEFAULT_SHEET_NAME in work_book.get_sheet_names():
        need_remove_sheet = work_book.get_sheet_by_name(settings.DEFAULT_SHEET_NAME)
        work_book.remove_sheet(need_remove_sheet)
    table = work_book.get_sheet_by_name(settings.OBS_ANONYMOUS_BUCKET_PAGE_NAME)
    table.delete_rows(1, 65536)
    table.append(settings.SCAN_OBS_EXCEL_BUCKET_TITLE)
    for bucket_list in anonymous_bucket_list:
        table.append(bucket_list)
    table = work_book.get_sheet_by_name(settings.OBS_SENSITIVE_FILE_PAGE_NAME)
    table.delete_rows(1, 65536)
    table.append(settings.SCAN_OBS_EXCEL_FILE_TITLE)
    for file_list in anonymous_file_list:
        table.append(file_list)
    buf = StringIO()
    work_book.save(buf)
    buf.seek(0)
    return buf.read()


def output_cla_excel(list_data):
    """output list data to excel"""
    work_book = openpyxl.Workbook()
    if settings.CLA_EXCEL_PAGE_NAME not in work_book.get_sheet_names():
        work_book.create_sheet(settings.CLA_EXCEL_PAGE_NAME)
    if settings.DEFAULT_SHEET_NAME in work_book.get_sheet_names():
        need_remove_sheet = work_book.get_sheet_by_name(settings.DEFAULT_SHEET_NAME)
        work_book.remove_sheet(need_remove_sheet)
    table = work_book.get_sheet_by_name(settings.CLA_EXCEL_PAGE_NAME)
    table.delete_rows(1, 65536)
    table.append(settings.CLA_EXCEL_TITLE)
    for bucket_dict in list_data:
        values = list(bucket_dict.values())
        values[6] = "{}%".format(values[6])
        values[7] = "{}%".format(values[7])
        table.append(values)
    buf = StringIO()
    work_book.save(buf)
    buf.seek(0)
    return buf.read()


def output_excel(list_data, page_name="Sheet1", title=None):
    """output list data to excel"""
    if title is None:
        title = list()
    work_book = openpyxl.Workbook()
    if page_name not in work_book.get_sheet_names():
        work_book.create_sheet(page_name)
    if settings.DEFAULT_SHEET_NAME in work_book.get_sheet_names():
        need_remove_sheet = work_book.get_sheet_by_name(settings.DEFAULT_SHEET_NAME)
        work_book.remove_sheet(need_remove_sheet)
    table = work_book.get_sheet_by_name(page_name)
    table.delete_rows(1, 65536)
    table.append(title)
    for dict_data in list_data:
        temp = dict_data.copy()
        del temp["id"]
        table.append(list(temp.values()))
    buf = StringIO()
    work_book.save(buf)
    buf.seek(0)
    return buf.read()


def output_table_excel(sheet_name, title_name, table_list_data):
    """output list data to excel"""
    work_book = openpyxl.Workbook()
    if sheet_name not in work_book.get_sheet_names():
        work_book.create_sheet(sheet_name)
    if settings.DEFAULT_SHEET_NAME in work_book.get_sheet_names():
        need_remove_sheet = work_book.get_sheet_by_name(settings.DEFAULT_SHEET_NAME)
        work_book.remove_sheet(need_remove_sheet)
    table = work_book.get_sheet_by_name(sheet_name)
    table.delete_rows(1, 65536)
    table.append(title_name)
    for table_list in table_list_data:
        table.append(table_list)
    buf = StringIO()
    work_book.save(buf)
    buf.seek(0)
    return buf.read()


def runserver_executor(func):
    """A decorator for a method to be executed only when the service is run"""
    @wraps(func)
    def wrapper(*args, **kw):
        if settings.IS_RUNSERVER:
            return func(*args, **kw)
    return wrapper


def list_param_check_and_trans(params, order_type=0, order_by="account"):
    """Handling list input parameters
    :param params : query param, ege: page, size
    :param order_type: 1 is descending, 2 is ascending
    :param order_by: the order_by
    :return: dict data, include: page, size, order_by, order_type, filter_name, filter_value
    """
    dict_data = dict()
    page, size = params.get("page", "1"), params.get("size", "100")
    # 1是降序， 0是升序
    order_type, order_by = params.get('order_type', order_type), params.get('order_by', order_by)
    if not page or not size:
        raise MgrException(ErrCode.STATUS_PARAMETER_ERROR)

    if not page.isdigit() or not size.isdigit():
        raise MgrException(ErrCode.STATUS_PARAMETER_ERROR)

    if int(page) < 1 or int(size) < 1:
        raise MgrException(ErrCode.STATUS_PARAMETER_ERROR)

    dict_data['page'], dict_data['size'] = int(page), int(size)
    dict_data['order_by'] = order_by
    if order_type:
        if not order_type.isdigit() or int(order_type) not in [0, 1]:
            raise MgrException(ErrCode.STATUS_PARAMETER_ERROR)
        dict_data['order_type'] = int(order_type)
    filter_name, filter_value = params.get("filter_name"), params.get("filter_value")
    if filter_name and filter_value:
        dict_data["filter_name"] = filter_name.strip()
        dict_data["filter_value"] = filter_value.strip()
    return dict_data


def get_max_page(total, size):
    """Get the maximum number of pages
    :param total: total page
    :param size: total size
    :return: max page size
    """
    if total == 0:
        return 1
    full_page_num, remain_num = divmod(total, size)
    if remain_num:
        max_page = full_page_num + 1
    else:
        max_page = full_page_num
    return max_page


def get_suitable_range(total, page, size):
    """Get the appropriate page number, corresponding to the slice of the range.
    :param total: the count of total
    :param page: the page of query
    :param size: the page size of query
    :return: page number, slice object
    """
    suitable_page = min(get_max_page(total, size), page)
    start = (suitable_page - 1) * size
    end = min(start + size, total)
    return suitable_page, slice(start, end)


def get_random_password(bit=12):
    """Automatically generate password"""
    words = string.ascii_lowercase + string.ascii_uppercase + string.digits
    list_word = random.sample(words, bit)
    list_word.extend(["!", "@", "#", "1", "a", "A"])
    return "".join(list_word)


def get_month_range(start_day, end_day):
    """Get the year, month, and day of a period of time"""
    if not isinstance(start_day, datetime.date):
        raise ValueError("start_day must be datetime.data")
    if not isinstance(end_day, datetime.date):
        raise ValueError("end_day must be datetime.data")
    months = (end_day.year-start_day.year)*12 + end_day.month - start_day.month
    month_range = list()
    for mon in range(start_day.month - 1, start_day.month + months):
        year_temp = start_day.year + mon // 12
        month_temp = mon % 12 + 1
        if len(str(month_temp)) == 2:
            value = "{}-{}".format(year_temp, month_temp)
        else:
            value = "{}-0{}".format(year_temp, month_temp)
        month_range.append(value)
    return month_range


def format_float(float_num):
    """format float eg: 26469833.59 to 26,469,833.59"""
    ret_list = list()
    int_num = int(float_num)
    str_num = str(int_num)
    len_str_num = len(str_num)
    for i in range(0, len_str_num):
        if (len_str_num - i) % 3 == 0 and len_str_num != len_str_num - i:
            ret_list.append(",")
            ret_list.append(str_num[i])
        else:
            ret_list.append(str_num[i])
    high_value = "".join(ret_list)
    lower_value = str(round(float_num - int_num, 2)).split(".")[-1]
    return high_value + "." + lower_value


class MgrException(Exception):
    """The Mgr Exception, use it to re"""
    def __init__(self, code, trans_para=None, trans_code=None, message=None, desc=None):
        self.code = code
        if isinstance(trans_para, str):
            trans_para = [trans_para]
        self.trans_para = trans_para
        self.trans_code = trans_code
        if desc is None:
            self.desc = translate_error_desc(code, [] if trans_para is None else trans_para)
        else:
            self.desc = desc
        super(MgrException, self).__init__(self.desc if message is None else message)


class BaseStatus:
    """eg: FREEZED = (0, "冻结")"""

    @classmethod
    def get_status_comment(cls):
        """get the dict eg: dict[0] = 冻结"""
        dict_data = dict()
        for attr, content in cls.__dict__.items():
            if attr.isupper() and isinstance(content, tuple):
                dict_data[content[0]] = content[1]
        return dict_data

    @classmethod
    def get_comment_status(cls):
        """get the dict eg: dict["冻结"] = 0"""
        dict_data = dict()
        for attr, content in cls.__dict__.items():
            if attr.isupper() and isinstance(content, tuple):
                dict_data[content[1]] = content[0]
        return dict_data
